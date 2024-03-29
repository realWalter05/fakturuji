from __future__ import print_function
from PIL import Image
from io import BytesIO
import os
from datetime import date
import openpyxl
import pandas as pd
import requests
import math
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.borders import Border, Side, BORDER_THICK, BORDER_THIN
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.worksheet.pagebreak import Break
from errno import EACCES, EPERM
import secrets

# Import module
import groupdocs_conversion_cloud
from shutil import copyfile


def fill_out_dodavatele(sheet, dodavatel_df, start_row):
    sheet["A" + str(start_row + 4)] = dodavatel_df.iloc[0]["dodavatel"]
    sheet["A" + str(start_row + 5)] = dodavatel_df.iloc[0]["ulice"]
    sheet["A" + str(start_row + 6)] = dodavatel_df.iloc[0]["mesto"]
    sheet["A" + str(start_row + 7)] = dodavatel_df.iloc[0]["zeme"]
    sheet["A" + str(start_row + 10)] = dodavatel_df.iloc[0]["zapis_rejstrik"]
    sheet["B" + str(start_row + 8)] = dodavatel_df.iloc[0]["ico"]
    sheet["B" + str(start_row + 9)] = dodavatel_df.iloc[0]["dic"]
    sheet["D" + str(start_row + 5)] = dodavatel_df.iloc[0]["telefon"]
    sheet["D" + str(start_row + 6)] = dodavatel_df.iloc[0]["email"]
    sheet["D" + str(start_row + 7)] = dodavatel_df.iloc[0]["web"]


def fill_out_odberatele(sheet, odberatel_df, start_row):
    sheet["H" + str(start_row + 4)] = odberatel_df.iloc[0]["odberatel"]
    sheet["H" + str(start_row + 5)] = odberatel_df.iloc[0]["ulice"]
    sheet["H" + str(start_row + 6)] = odberatel_df.iloc[0]["mesto"]
    sheet["H" + str(start_row + 7)] = odberatel_df.iloc[0]["zeme"]
    sheet["I" + str(start_row + 8)] = odberatel_df.iloc[0]["ico"]
    sheet["I" + str(start_row + 9)] = odberatel_df.iloc[0]["dic"]
    sheet["K" + str(start_row + 5)] = odberatel_df.iloc[0]["telefon"]
    sheet["K" + str(start_row + 6)] = odberatel_df.iloc[0]["email"]
    sheet["K" + str(start_row + 7)] = odberatel_df.iloc[0]["web"]


def get_mena_ending_reversed(mena):
    currencies = {
        "Kč": "CZK",
        "€" : "EUR",
        "$" : "USD",
        "£" : "GBP",
    }
    if mena in currencies:
        return currencies[mena]
    return "CZK"


def fill_out_extra_data(sheet, prenesena_dph, dodavatel_dph, descriptions, start_row):
    if dodavatel_dph:
        sheet["D" + str(start_row + 9)] = "Dodavatel je plátce DPH"
    if prenesena_dph:
        sheet["H" + str(start_row + 10)] = "Faktura vystavena v režimu přenesené daňové povinnosti. Daň odvede zákazník."
    if descriptions:
        sheet["A" + str(start_row + 20)] = descriptions


def fill_out_dates(sheet, dates, start_row):
    # Fill out dates values
    sheet["K" + str(start_row + 14)].value = dates["vystaveni_date"]
    sheet["K" + str(start_row + 15)].value = dates["zdanpl_date"]
    sheet["K" + str(start_row + 16)].value = dates["splatnost_date"]


def return_number(string_number):
    string_number = str(string_number).replace(" ", "")
    negative_number = False
    if "-" in string_number:
        string_number = string_number[1:]
        negative_number = True

    if "." in string_number.replace(",", "."):
        decimal = string_number.replace(",", ".")

        if decimal.replace(".", "").isnumeric():
            if negative_number:
                return float(decimal) * -1
            return float(decimal)

    if string_number.isnumeric():
        if negative_number:
            return int(string_number) * -1
        return int(string_number)
    return 0


def fill_out_items(sheet, items, start_row, description, mena):
    loop = 0
    for item in items:
        item_number = 21 + start_row + loop if description else 20 + start_row + loop

        if item["dodavka"]:
            sheet["A" + str(item_number)] = item["dodavka"]

        if item["pocet"]:
            sheet["E" + str(item_number)] = return_number(item["pocet"])
        if item["cena"]:
            sheet["F" + str(item_number)] = return_number(item["cena"])
        print(return_number(item["cena"]))

        if item["dph"]:
            sheet["H" + str(item_number)] = return_number(item["dph"])

        style_item(item_number, sheet, mena)

        loop = loop + 1


def get_dph_rates(sheet, items, start_row, items_count, description):
    # Get various dph rates
    dph_rates = {}
    for item in items:
        if not str(item["dph"]).replace(',', "").replace('.', "").isnumeric():
            continue
        dph = return_number(str(item["dph"]))
        if dph not in dph_rates:
            dph_rates[dph] = []

    # Save which items have which dph rate s
    for dph in dph_rates:
        for item in items:
            if not str(item["dph"]).replace(',', "").replace('.', "").isnumeric():
                continue
            d = return_number(str(item["dph"]))
            if d == dph:
                dph_rates[dph].append(items.index(item))

    # Write three dph rates to the sheet
    i = 3
    for dph in dph_rates:
        if i <= 0:
            break
        sheet["I" + str(start_row + i)].value = "DPH " + str(dph) + "%"

        sum_string = ""
        for cell in dph_rates[dph]:
            if not sum_string:
                sum_string = "=K" + str(start_row - items_count + 2 + cell) + "- I" + str(start_row - items_count + 2 + cell) if description\
                    else "=K" + str(start_row - items_count + 1 + cell) + "- I" + str(start_row - items_count + 1 + cell)
                continue

            sum_string = sum_string + " + K" + str(start_row - items_count + 2 + cell) + " - I" + str(start_row - items_count + 2 + cell) if description else\
                sum_string + " + K" + str(start_row - items_count + 1 + cell) + " - I" + str(start_row - items_count + 1 + cell)

        sheet["K" + str(start_row + i)].value = sum_string
        i -= 1


def fill_out_account_info(sheet, dodavatel_df, start_row, faktura_variable, column):
    # Fill out account values
    if not pd.isnull(dodavatel_df.iloc[0]["cislo_uctu"]) and not pd.isnull(dodavatel_df.iloc[0]["kod_banky"]):
        sheet[column + str(start_row + 2)].value = dodavatel_df.iloc[0]["cislo_uctu"] + "/" + dodavatel_df.iloc[0]["kod_banky"]
    else:
        sheet[column + str(start_row + 2)].value = dodavatel_df.iloc[0]["cislo_uctu"]
    sheet[column + str(start_row + 3)].value = dodavatel_df.iloc[0]["swift"]
    sheet[column + str(start_row + 4)].value = dodavatel_df.iloc[0]["iban"]

    if pd.isnull(dodavatel_df.iloc[0]["var_cislo"]):
        if not pd.isnull(dodavatel_df.iloc[0]["cislo_uctu"]) and not pd.isnull(dodavatel_df.iloc[0]["kod_banky"]):
            sheet["G" + str(start_row + 2)].value = faktura_variable
    else:
        sheet["G" + str(start_row + 2)].value = dodavatel_df.iloc[0]["var_cislo"]
    sheet["G" + str(start_row + 3)].value = dodavatel_df.iloc[0]["konst_cislo"]


def write_qr_platba_code(sheet, start_row, account_number, bank_code, items, var_cislo, prenesena_dph, mena):
    ammount = 0
    if not prenesena_dph:
        for item in items:
            value = (return_number(item["cena"]) + (return_number(item["cena"]) * return_number(item["dph"]) * 0.01)) * return_number(item["pocet"])
            ammount += value
    else:
        for item in items:
            value = return_number(item["cena"]) * return_number(item["pocet"])
            ammount += value

    if isinstance(ammount, float):
        ammount = round(ammount, 2)

    # Writing the qr platba code
    response = requests.get("https://api.paylibo.com/paylibo/generator/czech/image?accountNumber="+str(int(account_number))+
                            "&bankCode="+str(int(bank_code))+"&amount="+str(ammount)+"&currency="+get_mena_ending_reversed(mena)+"&vs="+str(var_cislo)+"&size=200")
    img = openpyxl.drawing.image.Image(Image.open(BytesIO(response.content)))
    img.width = 120
    img.height = 120
    img.anchor = 'A' + str(start_row + 1)
    sheet.add_image(img)


def style_item(item_number, sheet, mena):
    # Style item
    white_bottom_border = Border(bottom=Side(border_style=BORDER_THICK, color='FFFFFF'))
    gray_fill = PatternFill(start_color='00F3F3F3', end_color='00F3F3F3', fill_type='solid')

    sheet.row_dimensions[item_number].height = 24
    sheet.merge_cells(start_row=item_number, start_column=1, end_row=item_number, end_column=4)
    sheet["A" + str(item_number)].alignment = Alignment(vertical="center", shrinkToFit=True)
    sheet["A" + str(item_number)].fill = gray_fill
    sheet["A" + str(item_number)].border = white_bottom_border
    sheet["B" + str(item_number)].border = white_bottom_border
    sheet["C" + str(item_number)].border = white_bottom_border
    sheet["D" + str(item_number)].border = white_bottom_border

    sheet["E" + str(item_number)].alignment = Alignment(vertical="center", horizontal="center", shrinkToFit=True)
    sheet["E" + str(item_number)].fill = gray_fill
    sheet["E" + str(item_number)].border = white_bottom_border

    sheet.merge_cells(start_row=item_number, start_column=6, end_row=item_number, end_column=7)
    sheet["F" + str(item_number)].alignment = Alignment(vertical="center", horizontal="right", shrinkToFit=True)
    sheet["F" + str(item_number)].number_format = f"#,##0.00 {mena}"
    sheet["F" + str(item_number)].fill = gray_fill
    sheet["F" + str(item_number)].border = white_bottom_border
    sheet["G" + str(item_number)].border = white_bottom_border

    sheet["H" + str(item_number)].alignment = Alignment(vertical="center", shrinkToFit=True)
    sheet["H" + str(item_number)].number_format = "#0\%"
    sheet["H" + str(item_number)].fill = gray_fill
    sheet["H" + str(item_number)].border = white_bottom_border

    sheet.merge_cells(start_row=item_number, start_column=9, end_row=item_number, end_column=10)
    sheet["I" + str(item_number)].alignment = Alignment(vertical="center", shrinkToFit=True)
    sheet["I" + str(item_number)].number_format = f"#,##0.00 {mena}"
    sheet["I" + str(item_number)].value = "=IF(AND(E" + str(item_number) + "=\"\",F" +\
                                          str(item_number) + "=\"\"),\"\",E" + str(item_number) + "*F" + str(item_number) + ")"
    sheet["I" + str(item_number)].fill = gray_fill
    sheet["I" + str(item_number)].border = white_bottom_border
    sheet["J" + str(item_number)].border = white_bottom_border

    sheet.merge_cells(start_row=item_number, start_column=11, end_row=item_number, end_column=12)
    sheet["K" + str(item_number)].alignment = Alignment(vertical="center", shrinkToFit=True)
    sheet["K" + str(item_number)].number_format = f"#,##0.00 {mena}"
    sheet["K" + str(item_number)].value = "=IF(I" + str(item_number) + "=\"\", \"\", (I" + str(item_number) +\
                                          "+(I" + str(item_number) + "/100)*H" + str(item_number) + "))"
    sheet["K" + str(item_number)].fill = gray_fill
    sheet["K" + str(item_number)].border = white_bottom_border
    sheet["L" + str(item_number)].border = white_bottom_border


def style_items_faktura(sheet, start_row, description, mena):
    # Setting blank line
    sheet.row_dimensions[start_row + 18].height = 16.5
    sheet.row_dimensions[start_row + 19].height = 6

    if description:
        sheet.row_dimensions[start_row + 20].height = 24.6
        sheet.merge_cells(start_row=start_row + 20, start_column=1, end_row=start_row + 20, end_column=12)
        sheet["A" + str(start_row + 20)].alignment = Alignment(vertical="bottom", horizontal="left", wrap_text=True)
        sheet["A" + str(start_row + 20)].font = Font(size=10)

    loop = 0
    while loop < 8:
        item_number = 21 + start_row + loop if description else 20 + start_row + loop
        style_item(item_number, sheet, mena)

        loop = loop + 1

    # Merging columns
    sheet.merge_cells(start_row=start_row + 18, start_column=1, end_row=start_row + 18, end_column=4)
    sheet.merge_cells(start_row=start_row + 18, start_column=6, end_row=start_row + 18, end_column=7)
    sheet.merge_cells(start_row=start_row + 18, start_column=9, end_row=start_row + 18, end_column=11)

    # Changing fonts and alignments
    sheet["A" + str(start_row + 18)].font = Font(size=10, bold=True, color="FFFFFFFF")
    sheet["A" + str(start_row + 18)].alignment = Alignment(vertical="center", horizontal="center")
    sheet["E" + str(start_row + 18)].font = Font(size=10, bold=True, color="FFFFFFFF")
    sheet["E" + str(start_row + 18)].alignment = Alignment(vertical="center", horizontal="center")
    sheet["F" + str(start_row + 18)].font = Font(size=10, bold=True, color="FFFFFFFF")
    sheet["F" + str(start_row + 18)].alignment = Alignment(vertical="center", horizontal="center")
    sheet["H" + str(start_row + 18)].font = Font(size=10, bold=True, color="FFFFFFFF")
    sheet["H" + str(start_row + 18)].alignment = Alignment(vertical="center", horizontal="center")
    sheet["I" + str(start_row + 18)].font = Font(size=10, bold=True, color="FFFFFFFF")
    sheet["I" + str(start_row + 18)].alignment = Alignment(vertical="center", horizontal="center")
    sheet["L" + str(start_row + 18)].font = Font(size=10, bold=True, color="FFFFFFFF")
    sheet["L" + str(start_row + 18)].alignment = Alignment(vertical="center", horizontal="center")

    # Fill out default values
    sheet["A" + str(start_row + 18)].value = "Označení dodávky"
    sheet["E" + str(start_row + 18)].value = "Množství"
    sheet["F" + str(start_row + 18)].value = "Za kus"
    sheet["H" + str(start_row + 18)].value = "DPH"
    sheet["I" + str(start_row + 18)].value = "Bez DPH"
    sheet["L" + str(start_row + 18)].value = "S DPH"

    # Coloring
    sheet["A" + str(start_row + 18)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["E" + str(start_row + 18)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["F" + str(start_row + 18)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["H" + str(start_row + 18)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["I" + str(start_row + 18)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["K" + str(start_row + 18)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["L" + str(start_row + 18)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')


def style_first_part_faktura(sheet, start_row, faktura_numbering, variable_data):
    # Changing the height of rows in the first section
    sheet.row_dimensions[start_row].height = 24.5
    sheet.row_dimensions[start_row + 1].height = 25.8
    sheet.row_dimensions[start_row + 2].height = 12.4
    sheet.row_dimensions[start_row + 3].height = 19.6
    sheet.row_dimensions[start_row + 4].height = 16.8
    sheet.row_dimensions[start_row + 5].height = 16.5
    sheet.row_dimensions[start_row + 6].height = 16.5
    sheet.row_dimensions[start_row + 7].height = 16.5
    sheet.row_dimensions[start_row + 8].height = 16.5
    sheet.row_dimensions[start_row + 9].height = 16.5
    sheet.row_dimensions[start_row + 10].height = 16.8
    sheet.row_dimensions[start_row + 11].height = 13.8
    sheet.row_dimensions[start_row + 12].height = 11.6
    sheet.row_dimensions[start_row + 13].height = 16.5
    sheet.row_dimensions[start_row + 14].height = 16.5
    sheet.row_dimensions[start_row + 15].height = 16.5
    sheet.row_dimensions[start_row + 16].height = 16.5
    sheet.row_dimensions[start_row + 17].height = 8.4

    # Changing the width of columns
    sheet.column_dimensions['A'].width = 4.62
    sheet.column_dimensions['B'].width = 11.58
    sheet.column_dimensions['C'].width = 1.91
    sheet.column_dimensions['D'].width = 5.47
    sheet.column_dimensions['E'].width = 15.20
    sheet.column_dimensions['F'].width = 6.47
    sheet.column_dimensions['G'].width = 6.58
    sheet.column_dimensions['H'].width = 5.80
    sheet.column_dimensions['I'].width = 13.02
    sheet.column_dimensions['J'].width = 1.91
    sheet.column_dimensions['K'].width = 4.80
    sheet.column_dimensions['L'].width = 15.02

    # Merging columns in the first section
    sheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=5)

    sheet.merge_cells(start_row=start_row + 3, start_column=1, end_row=start_row + 3, end_column=5)
    sheet.merge_cells(start_row=start_row + 3, start_column=8, end_row=start_row + 3, end_column=12)

    sheet.merge_cells(start_row=start_row + 4, start_column=1, end_row=start_row + 4, end_column=5)
    sheet.merge_cells(start_row=start_row + 4, start_column=8, end_row=start_row + 4, end_column=12)

    sheet.merge_cells(start_row=start_row + 5, start_column=1, end_row=start_row + 5, end_column=2)
    sheet.merge_cells(start_row=start_row + 5, start_column=4, end_row=start_row + 5, end_column=5)
    sheet.merge_cells(start_row=start_row + 5, start_column=8, end_row=start_row + 5, end_column=9)
    sheet.merge_cells(start_row=start_row + 5, start_column=11, end_row=start_row + 5, end_column=12)

    sheet.merge_cells(start_row=start_row + 6, start_column=1, end_row=start_row + 6, end_column=2)
    sheet.merge_cells(start_row=start_row + 6, start_column=4, end_row=start_row + 6, end_column=5)
    sheet.merge_cells(start_row=start_row + 6, start_column=8, end_row=start_row + 6, end_column=9)
    sheet.merge_cells(start_row=start_row + 6, start_column=11, end_row=start_row + 6, end_column=12)

    sheet.merge_cells(start_row=start_row + 7, start_column=1, end_row=start_row + 7, end_column=2)
    sheet.merge_cells(start_row=start_row + 7, start_column=4, end_row=start_row + 7, end_column=5)
    sheet.merge_cells(start_row=start_row + 7, start_column=8, end_row=start_row + 7, end_column=9)
    sheet.merge_cells(start_row=start_row + 7, start_column=11, end_row=start_row + 7, end_column=12)

    sheet.merge_cells(start_row=start_row + 9, start_column=4, end_row=start_row + 9, end_column=5)
    sheet.merge_cells(start_row=start_row + 9, start_column=11, end_row=start_row + 9, end_column=12)

    sheet.merge_cells(start_row=start_row + 10, start_column=1, end_row=start_row + 11, end_column=5)
    sheet.merge_cells(start_row=start_row + 10, start_column=8, end_row=start_row + 11, end_column=12)

    sheet.merge_cells(start_row=start_row + 13, start_column=1, end_row=start_row + 13, end_column=2)
    sheet.merge_cells(start_row=start_row + 13, start_column=4, end_row=start_row + 13, end_column=5)
    sheet.merge_cells(start_row=start_row + 13, start_column=8, end_row=start_row + 13, end_column=9)
    sheet.merge_cells(start_row=start_row + 13, start_column=11, end_row=start_row + 13, end_column=12)

    sheet.merge_cells(start_row=start_row + 14, start_column=1, end_row=start_row + 14, end_column=2)
    sheet.merge_cells(start_row=start_row + 14, start_column=4, end_row=start_row + 14, end_column=5)
    sheet.merge_cells(start_row=start_row + 14, start_column=8, end_row=start_row + 14, end_column=9)
    sheet.merge_cells(start_row=start_row + 14, start_column=11, end_row=start_row + 14, end_column=12)

    sheet.merge_cells(start_row=start_row + 15, start_column=1, end_row=start_row + 15, end_column=2)
    sheet.merge_cells(start_row=start_row + 15, start_column=4, end_row=start_row + 15, end_column=5)
    sheet.merge_cells(start_row=start_row + 15, start_column=8, end_row=start_row + 15, end_column=9)
    sheet.merge_cells(start_row=start_row + 15, start_column=11, end_row=start_row + 15, end_column=12)

    sheet.merge_cells(start_row=start_row + 16, start_column=1, end_row=start_row + 16, end_column=2)
    sheet.merge_cells(start_row=start_row + 16, start_column=4, end_row=start_row + 16, end_column=5)
    sheet.merge_cells(start_row=start_row + 16, start_column=8, end_row=start_row + 16, end_column=9)
    sheet.merge_cells(start_row=start_row + 16, start_column=11, end_row=start_row + 16, end_column=12)

    # Fill out default values
    sheet["A" + str(start_row + 1)].value = "Faktura - daňový doklad"

    sheet["A" + str(start_row + 3)].value = "Dodavatel"
    sheet["A" + str(start_row + 8)].value = "IČ:"
    sheet["A" + str(start_row + 9)].value = "DIČ:"
    sheet["A" + str(start_row + 13)].value = variable_data[0][0]
    sheet["D" + str(start_row + 13)].value = variable_data[0][1]
    sheet["A" + str(start_row + 14)].value = variable_data[1][0]
    sheet["D" + str(start_row + 14)].value = variable_data[1][1]
    sheet["A" + str(start_row + 15)].value = variable_data[2][0]
    sheet["D" + str(start_row + 15)].value = variable_data[2][1]
    sheet["A" + str(start_row + 16)].value = variable_data[3][0]
    sheet["D" + str(start_row + 16)].value = variable_data[3][1]

    sheet["H" + str(start_row + 3)].value = "Odběratel"
    sheet["H" + str(start_row + 8)].value = "IČ:"
    sheet["H" + str(start_row + 9)].value = "DIČ:"
    sheet["H" + str(start_row + 14)].value = "Datum vystavení:"
    sheet["H" + str(start_row + 15)].value = "Datum zdan. plnění:"
    sheet["H" + str(start_row + 16)].value = "Datum splatnosti:"

    sheet["L" + str(start_row + 1)].value = faktura_numbering

    # Changing fonts and alignments
    sheet["A" + str(start_row + 1)].font = Font(size=14, bold=True)
    sheet["A" + str(start_row + 3)].font = Font(size=12, bold=True)
    sheet["A" + str(start_row + 4)].alignment = Alignment(shrinkToFit=True)
    sheet["A" + str(start_row + 5)].alignment = Alignment(shrinkToFit=True)
    sheet["A" + str(start_row + 6)].alignment = Alignment(shrinkToFit=True)
    sheet["A" + str(start_row + 7)].alignment = Alignment(shrinkToFit=True)
    sheet["A" + str(start_row + 10)].font = Font(size=9)
    sheet["A" + str(start_row + 10)].alignment = Alignment(vertical="center", wrap_text=True, shrinkToFit=True)

    #sheet["B" + str(start_row + 7)].alignment = Alignment(horizontal="left")
    sheet["B" + str(start_row + 8)].alignment = Alignment(horizontal="left", shrinkToFit=True)
    sheet["A" + str(start_row + 9)].alignment = Alignment(horizontal="left", shrinkToFit=True)

    sheet["D" + str(start_row + 5)].alignment = Alignment(horizontal="right", shrinkToFit=True)
    sheet["D" + str(start_row + 6)].alignment = Alignment(horizontal="right", shrinkToFit=True)
    sheet["D" + str(start_row + 7)].alignment = Alignment(horizontal="right", shrinkToFit=True)
    sheet["D" + str(start_row + 9)].font = Font(size=10)
    sheet["D" + str(start_row + 13)].alignment = Alignment(horizontal="right")
    sheet["D" + str(start_row + 14)].alignment = Alignment(horizontal="right")
    sheet["D" + str(start_row + 15)].alignment = Alignment(horizontal="right")
    sheet["D" + str(start_row + 16)].alignment = Alignment(horizontal="right")

    sheet["I" + str(start_row + 8)].alignment = Alignment(horizontal="left", shrinkToFit=True)
    sheet["I" + str(start_row + 9)].alignment = Alignment(horizontal="left", shrinkToFit=True)

    sheet["H" + str(start_row + 3)].font = Font(size=12, bold=True)
    sheet["H" + str(start_row + 4)].alignment = Alignment(shrinkToFit=True)
    sheet["H" + str(start_row + 5)].alignment = Alignment(shrinkToFit=True)
    sheet["H" + str(start_row + 6)].alignment = Alignment(shrinkToFit=True)
    sheet["H" + str(start_row + 7)].alignment = Alignment(shrinkToFit=True)
    sheet["H" + str(start_row + 10)].font = Font(size=9, bold=True)
    sheet["H" + str(start_row + 10)].alignment = Alignment(vertical="center", wrap_text=True)

    sheet["K" + str(start_row + 5)].alignment = Alignment(horizontal="right")
    sheet["K" + str(start_row + 6)].alignment = Alignment(horizontal="right")
    sheet["K" + str(start_row + 7)].alignment = Alignment(horizontal="right")

    sheet["K" + str(start_row + 13)].alignment = Alignment(horizontal="right")
    sheet["K" + str(start_row + 14)].alignment = Alignment(horizontal="right")
    sheet["K" + str(start_row + 15)].alignment = Alignment(horizontal="right")
    sheet["K" + str(start_row + 16)].alignment = Alignment(horizontal="right")

    sheet["L" + str(start_row + 1)].font = Font(size=14)
    sheet["L" + str(start_row + 1)].alignment = Alignment(horizontal="right")

    # Color first row
    sheet["A" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["B" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["C" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["D" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["E" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["F" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["G" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["H" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["I" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["J" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["K" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')
    sheet["L" + str(start_row)].fill = PatternFill(start_color='001383DD', end_color='001383DD', fill_type='solid')


def style_second_part_faktura(sheet, start_row, items_count, items, qr_platba, description, prenesena_dph, vystavila_osoba, mena):
    # Changing the height of rows in the second section
    sheet.row_dimensions[start_row].height = 12.5
    sheet.row_dimensions[start_row + 1].height = 18.8
    sheet.row_dimensions[start_row + 2].height = 19.5
    sheet.row_dimensions[start_row + 3].height = 18
    sheet.row_dimensions[start_row + 4].height = 18.8
    sheet.row_dimensions[start_row + 5].height = 9
    sheet.row_dimensions[start_row + 6].height = 9
    sheet.row_dimensions[start_row + 7].height = 12.6
    sheet.row_dimensions[start_row + 8].height = 19.5
    sheet.row_dimensions[start_row + 9].height = 3.6
    sheet.row_dimensions[start_row + 10].height = 17.3
    sheet.row_dimensions[start_row + 11].height = 24
    sheet.row_dimensions[start_row + 12].height = 25.1

    if not qr_platba:
        sheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=5)

        sheet.merge_cells(start_row=start_row + 2, start_column=4, end_row=start_row + 2, end_column=5)
        sheet.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 2, end_column=2)

        sheet.merge_cells(start_row=start_row + 3, start_column=1, end_row=start_row + 3, end_column=2)
        sheet.merge_cells(start_row=start_row + 3, start_column=4, end_row=start_row + 3, end_column=5)

        sheet.merge_cells(start_row=start_row + 4, start_column=1, end_row=start_row + 4, end_column=2)
        sheet.merge_cells(start_row=start_row + 4, start_column=4, end_row=start_row + 4, end_column=8)

        sheet["A" + str(start_row + 1)].font = Font(size=10, bold=True)
        sheet["A" + str(start_row + 1)].alignment = Alignment(vertical="center")

        sheet["A" + str(start_row + 2)].alignment = Alignment(vertical="center")
        sheet["A" + str(start_row + 3)].alignment = Alignment(vertical="center")
        sheet["A" + str(start_row + 4)].alignment = Alignment(vertical="center")

        sheet["D" + str(start_row + 2)].alignment = Alignment(shrinkToFit=True)
        sheet["D" + str(start_row + 3)].alignment = Alignment(shrinkToFit=True)
        sheet["D" + str(start_row + 4)].alignment = Alignment(shrinkToFit=True)

        sheet["F" + str(start_row + 1)].alignment = Alignment(vertical="center")
        sheet["F" + str(start_row + 2)].alignment = Alignment(vertical="center")

        sheet["A" + str(start_row + 1)].value = "Bankovní účet"
        sheet["A" + str(start_row + 2)].value = "Číslo účtu:"
        sheet["A" + str(start_row + 3)].value = "SWIFT:"
        sheet["A" + str(start_row + 4)].value = "IBAN:"

    else:
        sheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 6, end_column=2)
        sheet.merge_cells(start_row=start_row + 1, start_column=3, end_row=start_row + 1, end_column=5)

        sheet.merge_cells(start_row=start_row + 2, start_column=3, end_row=start_row + 2, end_column=4)

        sheet.merge_cells(start_row=start_row + 3, start_column=3, end_row=start_row + 3, end_column=4)

        sheet.merge_cells(start_row=start_row + 4, start_column=3, end_row=start_row + 4, end_column=4)
        sheet.merge_cells(start_row=start_row + 4, start_column=5, end_row=start_row + 4, end_column=8)

        sheet["C" + str(start_row + 1)].value = "Bankovní účet"
        sheet["C" + str(start_row + 2)].value = "Účet:"
        sheet["C" + str(start_row + 3)].value = "SWIFT:"
        sheet["C" + str(start_row + 4)].value = "IBAN:"

        sheet["C" + str(start_row + 1)].font = Font(size=10, bold=True)
        sheet["C" + str(start_row + 1)].alignment = Alignment(vertical="center")

        sheet["E" + str(start_row + 2)].alignment = Alignment(horizontal="left", shrinkToFit=True)
        sheet["E" + str(start_row + 2)].font = Font(size=10)
        sheet["E" + str(start_row + 3)].alignment = Alignment(horizontal="left", shrinkToFit=True)
        sheet["E" + str(start_row + 4)].alignment = Alignment(horizontal="left", shrinkToFit=True)

    # Merging columns in the second section
    sheet.merge_cells(start_row=start_row + 1, start_column=6, end_row=start_row + 1, end_column=8)
    sheet.merge_cells(start_row=start_row + 1, start_column=7, end_row=start_row + 1, end_column=8)
    sheet.merge_cells(start_row=start_row + 1, start_column=11, end_row=start_row + 1, end_column=12)

    sheet.merge_cells(start_row=start_row + 2, start_column=7, end_row=start_row + 2, end_column=8)
    sheet.merge_cells(start_row=start_row + 2, start_column=11, end_row=start_row + 2, end_column=12)

    sheet.merge_cells(start_row=start_row + 3, start_column=7, end_row=start_row + 3, end_column=8)
    sheet.merge_cells(start_row=start_row + 3, start_column=11, end_row=start_row + 3, end_column=12)

    sheet.merge_cells(start_row=start_row + 4, start_column=11, end_row=start_row + 4, end_column=12)

    sheet.merge_cells(start_row=start_row + 7, start_column=9, end_row=start_row + 7, end_column=12)

    sheet.merge_cells(start_row=start_row + 8, start_column=9, end_row=start_row + 8, end_column=12)

    sheet.merge_cells(start_row=start_row + 10, start_column=1, end_row=start_row + 10, end_column=12)

    sheet.merge_cells(start_row=start_row + 11, start_column=1, end_row=start_row + 11, end_column=12)

    sheet.merge_cells(start_row=start_row + 12, start_column=1, end_row=start_row + 12, end_column=2)

    # Fill out default values
    sheet["A" + str(start_row + 10)].value = "Dovolujeme si vás upozornit, že v případě nedodržení data splatnosti Vám můžeme účtovat zákonný úrok z prodlení."
    sheet["A" + str(start_row + 11)].value = vystavila_osoba
    sheet["A" + str(start_row + 12)].value = "Převzal:"

    # Add page break
    sheet.row_breaks.append(Break(start_row + 12))

    # Get and write dph rates
    if not prenesena_dph:
        get_dph_rates(sheet, items, start_row, items_count, description)

    sheet["E" + str(start_row + 12)].value = "dne"

    sheet["F" + str(start_row + 1)].value = "Symbol:"
    sheet["F" + str(start_row + 2)].value = "var:"
    sheet["F" + str(start_row + 3)].value = "konst:"

    sheet["I" + str(start_row + 4)].value = "BEZ DPH"
    sheet["I" + str(start_row + 7)].value = "CELKEM S DPH" if not prenesena_dph else "Celkem"

    # Changing fonts and alignments
    sheet["A" + str(start_row + 10)].font = Font(size=9)
    sheet["A" + str(start_row + 10)].alignment = Alignment(vertical="center")
    sheet["A" + str(start_row + 11)].alignment = Alignment(vertical="center")
    sheet["A" + str(start_row + 12)].alignment = Alignment(horizontal="left")

    sheet["D" + str(start_row + 2)].alignment = Alignment(horizontal="left")
    sheet["D" + str(start_row + 3)].alignment = Alignment(horizontal="left")
    sheet["D" + str(start_row + 4)].alignment = Alignment(horizontal="left")

    sheet["G" + str(start_row + 2)].alignment = Alignment(horizontal="left", shrinkToFit=True)
    sheet["G" + str(start_row + 3)].alignment = Alignment(horizontal="left", shrinkToFit=True)

    sheet["F" + str(start_row + 1)].font = Font(size=10, bold=True)

    sheet["I" + str(start_row + 1)].font = Font(size=7, bold=True, color="00636363")
    sheet["I" + str(start_row + 1)].alignment = Alignment(vertical="center", horizontal="right")
    sheet["I" + str(start_row + 2)].font = Font(size=7, bold=True, color="00636363")
    sheet["I" + str(start_row + 2)].alignment = Alignment(vertical="center", horizontal="right")
    sheet["I" + str(start_row + 3)].font = Font(size=7, bold=True, color="00636363")
    sheet["I" + str(start_row + 3)].alignment = Alignment(vertical="center", horizontal="right")
    sheet["I" + str(start_row + 4)].font = Font(size=7, bold=True, color="00636363")
    sheet["I" + str(start_row + 4)].alignment = Alignment(vertical="center", horizontal="right")

    sheet["K" + str(start_row + 1)].alignment = Alignment(vertical="center", horizontal="right", shrinkToFit=True)
    sheet["K" + str(start_row + 1)].number_format = f"#,##0.00 {mena}"

    sheet["K" + str(start_row + 2)].alignment = Alignment(vertical="center", horizontal="right", shrinkToFit=True)
    sheet["K" + str(start_row + 2)].number_format = f"#,##0.00 {mena}"

    sheet["K" + str(start_row + 3)].alignment = Alignment(vertical="center", horizontal="right", shrinkToFit=True)
    sheet["K" + str(start_row + 3)].number_format = f"#,##0.00 {mena}"

    sheet["K" + str(start_row + 4)].alignment = Alignment(vertical="center", horizontal="right", shrinkToFit=True)
    sheet["K" + str(start_row + 4)].number_format = f"#,##0.00 {mena}"
    sheet["K" + str(start_row + 4)].value = "=SUM(I"+str(start_row - items_count + 2)+":I"+str(start_row)+")" if description else \
        "=SUM(I"+str(start_row - items_count + 1)+":I"+str(start_row-1)+")"

    sheet["I" + str(start_row + 5)].border = Border(bottom=Side(border_style=BORDER_THIN, color='000000'))
    sheet["J" + str(start_row + 5)].border = Border(bottom=Side(border_style=BORDER_THIN, color='000000'))
    sheet["K" + str(start_row + 5)].border = Border(bottom=Side(border_style=BORDER_THIN, color='000000'))
    sheet["L" + str(start_row + 5)].border = Border(bottom=Side(border_style=BORDER_THIN, color='000000'))

    sheet["I" + str(start_row + 7)].font = Font(size=9, bold=True, color="00636363")
    sheet["I" + str(start_row + 7)].alignment = Alignment(vertical="center", horizontal="right")

    sheet["I" + str(start_row + 8)].font = Font(size=15)
    sheet["I" + str(start_row + 8)].alignment = Alignment(vertical="center", horizontal="right", shrinkToFit=True)
    sheet["I" + str(start_row + 8)].number_format = f"#,##0.00 {mena}"
    sheet["I" + str(start_row + 8)].value = "=K"+str(start_row + 4)+"+K"+str(start_row + 3)+"+K"+str(start_row + 2)+"+K"+str(start_row + 1)

    # Coloring rows
    for i in range(9):
        sheet["A" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        sheet["B" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        sheet["C" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        sheet["D" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        sheet["E" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        sheet["F" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        sheet["G" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        sheet["H" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        sheet["I" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        sheet["J" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        sheet["K" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')
        sheet["L" + str(start_row + i)].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type='solid')


def set_print_settings(sheet):
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToHeight = False
    sheet.page_margins.left = 0.7
    sheet.page_margins.right = 0.7
    sheet.page_margins.top = 0.75
    sheet.page_margins.bottom = 0.75
    sheet.page_margins.header = 0.3
    sheet.page_margins.footer = 0.3


def create_faktura(sheet, start_row, items, faktura_numbering, dodavatel_df, mena, qr_platba, dates, prenesena_dph, dodavatel_dph, descriptions, vystavila_osoba, variable_data):
    set_print_settings(sheet)
    style_first_part_faktura(sheet, start_row, faktura_numbering, variable_data)
    style_items_faktura(sheet, start_row, descriptions, mena)
    fill_out_dates(sheet, dates, start_row)
    fill_out_extra_data(sheet, prenesena_dph, dodavatel_dph, descriptions, start_row)

    second_start_row = 9
    if len(items) > 8:
        second_start_row = len(items) + 1

    style_second_part_faktura(sheet, second_start_row + start_row + 19, second_start_row, items, qr_platba, descriptions, prenesena_dph, vystavila_osoba, mena)
    c = "D"
    if qr_platba:
        c = "E"
        try:
            write_qr_platba_code(sheet, second_start_row + start_row + 19, dodavatel_df.iloc[0]["cislo_uctu"], dodavatel_df.iloc[0]["kod_banky"], items,
                                 faktura_numbering, prenesena_dph, mena)
        except Exception as e:
            print(e)
            c = "D"

    fill_out_account_info(sheet, dodavatel_df, second_start_row + start_row + 19, faktura_numbering, c)


def find_start_row(sheet):
    final_end_msg = ""
    for row in sheet.rows:
        if row[0].value == "Dovolujeme si vás upozornit, že v případě nedodržení data splatnosti Vám můžeme účtovat zákonný úrok z prodlení.":
            final_end_msg = row[0]

    return final_end_msg.row + 3


def get_faktura_number(sheet):
    final_faktura_title = ""
    for row in sheet.rows:
        if row[0].value == "Faktura - daňový doklad":
            final_faktura_title = row[0]

    last_number = sheet["L" + str(final_faktura_title.row)].value

    # Get the number without year
    seperator_number = len(str(last_number)) - 3

    faktura_number_second_part = str(last_number)[seperator_number:]
    faktura_number_first_part = str(last_number)[:seperator_number]
    formatted_second_part_number = "{:03d}".format(int(faktura_number_second_part) + 1)

    new_faktura_number = faktura_number_first_part + formatted_second_part_number
    return new_faktura_number


class ExcelWriter:

    def __init__(self):
        # Creates the workbook
        self.wb = openpyxl.Workbook()
        self.delete_empty_sheet()

        self.dodavatel = None
        self.odberatel = None

        self.odberatele = None
        self.dodavatele = None

        self.status = None
        self.faktura_numbering = None
        self.default_sheetnames = self.wb.sheetnames
        self.sheet_index = None
        self.sheet_max_row = None
        self.start_row = None
        self.dodavatel_df = None
        self.sheet = None


    def create_faktura(self, dodavatel_list, odberatel_list, items, prenesena_dph, dodavatel_dph, mena, qr_platba, dates, descriptions, def_faktura_numbering, vystavila_osoba, variable_data):
        # Get odberatel and dodavatel names
        if dodavatel_list:
            self.dodavatel = dodavatel_list[0]

            # Convert them to pandas
            self.dodavatele = pd.DataFrame([dodavatel_list], columns=["dodavatel","ulice","mesto","zeme","ico","dic","zapis_rejstrik","telefon","email","web","cislo_uctu","kod_banky","iban","swift","var_cislo","konst_cislo"])

            # Set rows with dodavatel/odberatel in pandas dataframe
            self.dodavatel_df = self.dodavatele[(self.dodavatele["dodavatel"] == self.dodavatel)]

        if odberatel_list:
            self.odberatel = odberatel_list[0]

            # Convert them to pandas
            self.odberatele = pd.DataFrame([odberatel_list], columns=["odberatel","ulice","mesto","zeme","ico","dic","zapis_rejstrik","telefon","email","web"])

            # Set rows with dodavatel/odberatel in pandas dataframe
            self.odberatel_df = self.odberatele[(self.odberatele["odberatel"] == self.odberatel)]

        self.set_start_row()
        self.set_faktura_numbering(def_faktura_numbering)
        self.set_sheet()

        try:
            # Create and fill out the faktura template
            create_faktura(self.sheet, self.start_row, items, self.faktura_numbering, self.dodavatel_df, mena, qr_platba, dates, prenesena_dph, dodavatel_dph, descriptions, vystavila_osoba, variable_data)
        except Exception as e:
            self.status = "errQrPlatba"
            print(f"err qr platba {e}")
            return

        # Fill dodavatel
        fill_out_dodavatele(self.sheet, self.dodavatel_df, self.start_row)

        # Fills odberatel
        fill_out_odberatele(self.sheet, self.odberatel_df, self.start_row)

        # Fill out items
        if len(items) > 0:
            fill_out_items(self.sheet, items, self.start_row, descriptions, mena)

    def set_start_row(self):
        self.start_row = 1
        if self.dodavatel in self.wb:
            # Dodavatel is already in excel
            self.start_row = find_start_row(self.sheet)


    def set_faktura_numbering(self, faktura_numbering):
        if faktura_numbering:
            # Set by user
            self.faktura_numbering = faktura_numbering

        elif self.dodavatel in self.wb:
            # Already existing
            self.faktura_numbering = get_faktura_number(self.sheet)

        else:
            # Setting default by us
            self.faktura_numbering = str(date.today().year)+"001"


    def set_sheet(self):
        if self.dodavatel in self.wb:
            self.sheet = self.wb[self.dodavatel]
            return
        self.sheet = self.wb.create_sheet(self.dodavatel_df.iloc[0]["dodavatel"])


    def delete_empty_sheet(self):
        # Deleting the default sheet if empty
        if self.wb["Sheet"]:
            if not len(list(self.wb["Sheet"].rows)) and not len(list(self.wb["Sheet"].columns)):
                self.wb.remove(self.wb["Sheet"])


    def save_faktura_in_excel(self):
        # Saving the new faktura
        self.wb.save("faktura" + secrets.token_hex(4) + ".xlsx")


    def get_virtual_save(self):
        return save_virtual_workbook(self.wb)